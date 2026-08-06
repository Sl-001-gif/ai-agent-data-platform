# -*- coding: utf-8 -*-
"""统计类数据结构化采集管道（stat-data-pipeline）。

三阶段：
  1. 详情页采集：从 gov_info_record 取 统计月报/统计公报/统计分析 三类记录，
     抓详情页正文全文（去导航/页眉/分享噪音）与附件（.xlsx/.xls）元数据 -> stat_doc
  2. 月报卡解析：下载 xlsx 附件，openpyxl 自适应表头识别，
     产出「期间 x 指标 x 区县」结构化指标 -> stat_indicator（价值最高）
  3. 正文指标抽取：公报/分析正文，规则正则优先，AI_API_KEY 存在时 LLM 兜底 -> stat_indicator

幂等：stat_doc 按 gov_record_id 唯一键 upsert；stat_indicator 按唯一键 INSERT IGNORE；
parse_status 断点续爬（PENDING/DONE/TEXT_DONE/XLSX_DONE/XLSX_FAIL/FAILED）。
复用 tools/scraper/gov_scraper.py 的 make_session/fetch/normalize_text，不改其行为。

用法：
    python stat_scraper.py --stage 1                 # 详情页采集（478 条）
    python stat_scraper.py --stage 2                 # 月报卡 xlsx 解析（85 份）
    python stat_scraper.py --stage 3                 # 公报/分析正文指标抽取
    python stat_scraper.py --stage all --limit 10    # 三阶段小批量调试
    python stat_scraper.py --stage 2 --force         # 重跑已解析文档
    python stat_scraper.py --stage 2 --dry-run       # 只打印解析样例不落库
"""
import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from io import BytesIO
from urllib.parse import urljoin

import pymysql

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

from bs4 import BeautifulSoup

# 复用既有爬虫工具（只 import 不改动）
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from gov_scraper import fetch, make_session, normalize_text  # noqa: E402

DB = dict(host="localhost", port=3306, user="root", password="Admin@123456",
          database="ai_agent_data", charset="utf8mb4")

SITE = "https://www.shaoyang.gov.cn"
INTERVAL = 1.0            # 合规限流：1s/条
TIMEOUT = 30
CATEGORY_MAP = {"tjyb": "统计月报", "stjgb": "统计公报", "stjfx": "统计分析"}

# 区县短名 -> 官方全名（月报卡列头用短名，统一归一化）
REGION_MAP = {
    "全市": "全市", "双清": "双清区", "大祥": "大祥区", "北塔": "北塔区",
    "新邵": "新邵县", "邵阳县": "邵阳县", "隆回": "隆回县", "洞口": "洞口县",
    "绥宁": "绥宁县", "新宁": "新宁县", "城步": "城步苗族自治县",
    "武冈": "武冈市", "邵东": "邵东市",
}

# sheet 表头定位：首列精确轴词（指标/项目/行业/地区…）+ 其余列关键词（数值/增速/排名…）
_AXIS_HEADERS = ("指标", "项目", "行业", "地区", "年份", "月份", "主要指标")
_HEADER_DETECT = ("绝对额", "增速", "增长", "增减", "同比", "排名", "位次", "产销率", "余额", "累计")
_RANK_KW = ("排名", "位次", "名次")
_GROWTH_KW = ("增速", "增幅", "增长率", "涨幅", "同比", "环比", "增减", "增长", "下降")
_VALUE_KW = ("绝对额", "总额", "总量", "余额", "产值", "收入", "支出", "完成",
             "销售", "产量", "增加值", "投资", "零售", "消费", "用电", "人口",
             "面积", "数量", "金额", "规模", "产销率")
_SHEET_SKIP = ("目录", "封面")

# 正文噪音行
_NOISE_RE = re.compile(r"扫一扫|打开当前页|上一篇|下一篇|相关文章|打印|字号|分享到|"
                       r"复制|纠错|版权所有|主办单位|承办单位|备案号|网站地图|无障碍|长者版")

# 阶段 3 已知指标规则（别名正则 -> 规范名）
_KNOWN_RULES = [
    (r"地区生产总值(?:（GDP）|\(GDP\))?", "地区生产总值（GDP）"),
    (r"第一产业(?:完成)?增加值", "第一产业增加值"),
    (r"第二产业(?:完成)?增加值", "第二产业增加值"),
    (r"第三产业(?:完成)?增加值", "第三产业增加值"),
    (r"规模以上工业增加值|规模工业增加值", "规模工业增加值"),
    (r"固定资产投资(?:（不含农户）)?", "固定资产投资"),
    (r"社会消费品零售总额", "社会消费品零售总额"),
    (r"地方一般公共预算收入|一般公共预算收入", "一般公共预算收入"),
    (r"进出口总额", "进出口总额"),
    (r"全体居民人均可支配收入|居民人均可支配收入", "居民人均可支配收入"),
    (r"城镇居民人均可支配收入", "城镇居民人均可支配收入"),
    (r"农村居民人均可支配收入", "农村居民人均可支配收入"),
    (r"居民消费价格(?:指数)?", "居民消费价格指数"),
]
_UNIT_RE = r"(亿元|万元|元|万人|万亩|万吨|万人次|亿千瓦时|亿千瓦小时|个百分点|%|户|家|个|亿元)"
_NUM_RE = r"(\d[\d,]*(?:\.\d+)?)"


# ============================================================
# 通用工具
# ============================================================
def _to_number(s):
    """数字文本 -> float；空/非数字返回 None（容忍千分位、%、全角逗号）。"""
    if s is None:
        return None
    t = str(s).strip().replace(",", "").replace("，", "").replace("%", "").replace(" ", "")
    if t in ("", "-", "--", "—", "//", "/"):
        return None
    try:
        return float(t)
    except ValueError:
        return None


def clean_indicator_name(name):
    """指标名清洗：去空白、去前导编号（一、/（一））。"""
    s = normalize_text(name or "")
    s = re.sub(r"^[一二三四五六七八九十]+、", "", s)
    s = re.sub(r"^[（(][一二三四五六七八九十]+[)）]", "", s)
    s = s.strip(" \t:：-—")
    return s


def normalize_period(p, year):
    """期间归一化：2025年1-9月 / 1-9月(补年份) / 2025年 / 1-12月。"""
    s = re.sub(r"\s+", "", p or "")
    m = re.match(r"^(20\d{2})?年?(\d{1,2}(?:-\d{1,2})?月)?$", s)
    if not m:
        return s or (f"{year}年" if year else "")
    y, mon = m.groups()
    if y:
        return s if "年" in s else f"{y}年{mon}"
    return f"{year}年{mon}" if mon else (f"{year}年" if year else s)


def infer_year(title, doc_date=None):
    m = re.search(r"(20\d{2})年", title or "")
    if m:
        return m.group(1)
    if doc_date:
        return str(doc_date)[:4]
    return str(datetime.now().year)


def period_from_title(title):
    m = re.search(r"(20\d{2})年(\d{1,2})-(\d{1,2})月", title or "")
    if m:
        return "%s年%s-%s月" % m.groups()
    m = re.search(r"(20\d{2})年", title or "")
    return "%s年" % m.group(1) if m else ""


def _unit_from_header(hdr):
    m = re.search(r"[（(]([^（）()]{1,12})[）)]", hdr)
    if m:
        u = m.group(1).strip()
        if re.search(r"[\u4e00-\u9fa5%]", u):
            return u
    return ""


# ============================================================
# 阶段 1：详情页正文与附件
# ============================================================
def extract_doc_content(html):
    """提取详情页正文（UCAPCONTENT 或 div.wenz 容器）与附件链接。

    返回 (正文文本, 附件相对链接或 None)。正文段落以换行分隔、行内文本保持连续
    （HTML 中数字被拆进多个 span，需无分隔符拼接后再按段落断开）。
    """
    m = re.search(r"<UCAPCONTENT>(.*?)</UCAPCONTENT>", html or "", re.S | re.I)
    inner = m.group(1) if m else None
    if not inner:
        m2 = re.search(r'<div[^>]*class="wenz"[^>]*>(.*?)</div>\s*</div>', html or "", re.S)
        inner = m2.group(1) if m2 else (html or "")
    attach = None
    am = re.search(r'href="([^"]+\.(?:xlsx|xls))"', inner, re.I)
    if am:
        attach = am.group(1)
    inner2 = re.sub(r"</p>|</div>|<br\s*/?>", "\n", inner)
    text = BeautifulSoup(inner2, "html.parser").get_text("")
    lines = [normalize_text(l) for l in text.split("\n")]
    lines = [l for l in lines if len(l) >= 4 and not _NOISE_RE.search(l)]
    return "\n".join(lines), attach


def fetch_docs(cur, categories, limit, force):
    marks = ",".join(["%s"] * len(categories))
    args = list(categories)
    if force:
        cur.execute(
            "SELECT id, category, title, publish_date, source_url FROM gov_info_record "
            "WHERE category IN (%s) ORDER BY id" % marks, args)
    else:
        cur.execute(
            "SELECT id, category, title, publish_date, source_url FROM gov_info_record "
            "WHERE category IN (%s) AND id NOT IN "
            "(SELECT gov_record_id FROM stat_doc WHERE parse_status IN "
            "('DONE','TEXT_DONE','XLSX_DONE','XLSX_FAIL')) ORDER BY id" % marks, args)
    rows = cur.fetchall()
    if limit:
        rows = rows[:limit]
    return rows


def stage1(conn, cur, session, categories, limit, force, dry_run, verbose):
    rows = fetch_docs(cur, categories, limit, force)
    ok = failed = skipped = 0
    for rid, cat, title, pub_date, url in rows:
        if not dry_run and not force:
            cur.execute("SELECT parse_status FROM stat_doc WHERE gov_record_id=%s", (rid,))
            row = cur.fetchone()
            if row and row[0] in ("DONE", "TEXT_DONE", "XLSX_DONE", "XLSX_FAIL"):
                skipped += 1
                continue
        try:
            html = fetch(session, url)
        except Exception as exc:
            failed += 1
            print("[详情失败] %s -> %s" % (url, exc), file=sys.stderr)
            if not dry_run:
                cur.execute(
                    "INSERT INTO stat_doc (gov_record_id, category, title, doc_date, source_url, parse_status, fail_reason) "
                    "VALUES (%s,%s,%s,%s,%s,'FAILED',%s) "
                    "ON DUPLICATE KEY UPDATE parse_status='FAILED', fail_reason=VALUES(fail_reason)",
                    (rid, cat, title, pub_date, url, str(exc)[:400]))
                conn.commit()
            continue
        text, attach = extract_doc_content(html)
        attach_url = urljoin(url, attach) if attach else None
        # 状态语义：月报有附件 -> DONE(待 xlsx 解析)；月报无附件 -> TEXT_DONE(无指标源)；
        # 公报/分析 -> DONE(待正文指标抽取，stage3 完成后置 TEXT_DONE)
        status = "DONE" if (attach or cat != "统计月报") else "TEXT_DONE"
        if dry_run:
            print("[dry] id=%s %s | 正文 %d 字 | 附件 %s | status=%s"
                  % (rid, title[:24], len(text or ""), attach_url, status))
        else:
            cur.execute(
                "INSERT INTO stat_doc (gov_record_id, category, title, doc_date, source_url, "
                "content, attachment_url, attachment_name, parse_status) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                "ON DUPLICATE KEY UPDATE content=VALUES(content), attachment_url=VALUES(attachment_url), "
                "attachment_name=VALUES(attachment_name), parse_status=VALUES(parse_status), fail_reason=NULL",
                (rid, cat, title, pub_date, url, text, attach_url,
                 os.path.basename(attach) if attach else None, status))
            conn.commit()
            ok += 1
            if verbose:
                print("[%s] id=%s %s -> %s 正文%d字 附件:%s"
                      % (cat, rid, title[:22], status, len(text or ""), attach_url or "-"))
        time.sleep(INTERVAL)
    print("stage1 完成：共 %d 条，成功 %d / 失败 %d / 跳过 %d" % (len(rows), ok, failed, skipped))


# ============================================================
# 阶段 2：xlsx 月报卡解析
# ============================================================
def parse_xlsx_sheet(ws, sheet, year):
    """解析单个 sheet -> 指标行列表（dict）。支持三种表形：
    A 区县表（首列=区县短名，标题行=指标名，含 绝对额/增速/排名 列）
    B 指标表 4 列（指  标 | 期间 | 绝对额(单位) | 增速(%)）
    C 指标表 3 列（指标/行业 | 期间 | 增速（%）等）
    """
    raw = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        raw.append(["" if c is None else str(c).strip() for c in row])
        if i >= 249:
            break
    raw = [r for r in raw if any(r)]
    if not raw:
        return []
    max_col = max(len(r) for r in raw)

    def cell(r, c):
        return raw[r][c] if r < len(raw) and c < len(raw[r]) else ""

    title = next((r[0] for r in raw if r[0]), "")
    # 表头行：前 12 行内命中表头关键词
    def _is_header_row(r):
        """表头行判定：首列为轴词（指标/项目/行业/地区/年份），或非首列含数值/增速/排名关键词。"""
        axis0 = re.sub(r"\s+", "", cell(r, 0))
        if axis0 in _AXIS_HEADERS:
            return True
        rest = re.sub(r"\s+", "", " ".join(cell(r, c) for c in range(1, max_col)))
        return (any(k in rest for k in _HEADER_DETECT)
                or any(k in rest for k in _RANK_KW)
                or any(k in rest for k in _GROWTH_KW))

    header_idx = [r for r in range(min(12, len(raw))) if _is_header_row(r)]
    if not header_idx:
        return []
    hdr_all = re.sub(r"\s+", "", " ".join(cell(r, c) for r in header_idx for c in range(max_col)))
    if "年份" in hdr_all or "月份" in hdr_all:
        return []  # 跳过 年份x月份 时间序列表（如湖南省财政对比表）
    # 列角色
    col_role, col_unit = {}, {}
    for c in range(1, max_col):
        hdr = re.sub(r"\s+", "", " ".join(cell(r, c) for r in header_idx if cell(r, c)))
        if not hdr:
            continue
        if any(k in hdr for k in _RANK_KW):
            col_role[c] = "rank"
        elif any(k in hdr for k in _GROWTH_KW) and not re.search(r"(额|余额)", hdr):
            col_role[c] = "growth"
            col_unit[c] = "个百分点" if "百分点" in hdr else "%"
        elif any(k in hdr for k in _VALUE_KW):
            col_role[c] = "value"
            col_unit[c] = _unit_from_header(hdr)
        else:
            col_role[c] = "value"
            col_unit[c] = _unit_from_header(hdr)
    # 期间
    period = ""
    for r in header_idx:
        for c in range(1, max_col):
            v = cell(r, c)
            if re.match(r"^(20\d{2})?年?\s*\d{1,2}-?\d*月", v) or re.match(r"^\d{1,2}月$", v):
                period = normalize_period(v, year)
                break
        if period:
            break
    if not period:
        period = "%s年" % year if year else ""
    vcol = next((c for c, role in col_role.items() if role == "value"), None)
    gcol = next((c for c, role in col_role.items() if role == "growth"), None)
    rcol = next((c for c, role in col_role.items() if role == "rank"), None)
    data_start = max(header_idx) + 1
    out = []
    for r in range(data_start, len(raw)):
        axis = cell(r, 0)
        if not axis:
            continue
        vals = {c: _to_number(cell(r, c)) for c in (vcol, gcol, rcol) if c is not None}
        if not any(v is not None for v in vals.values()):
            continue
        if axis in REGION_MAP:
            # A 区县表：indicator = sheet 标题，一行 = 一个区县
            base = dict(indicator_name=clean_indicator_name(title) or "统计指标", region=REGION_MAP[axis],
                        period=period, sheet_name=sheet, confidence="high",
                        generator_type="RULE", source_type="XLSX")
            rec = dict(base)
            if vcol is not None and vals.get(vcol) is not None:
                rec["value"], rec["unit"] = vals[vcol], col_unit.get(vcol)
            if gcol is not None and vals.get(gcol) is not None:
                rec["growth_rate"] = vals[gcol]
            if "value" in rec or "growth_rate" in rec:
                out.append(rec)
            if rcol is not None and vals.get(rcol) is not None:
                rank = dict(base)
                rank["indicator_name"] = (clean_indicator_name(title) or "统计指标") + "排名"
                rank["value"], rank["unit"] = vals[rcol], "名"
                out.append(rank)
        else:
            # B/C 指标表：indicator = 首列，region = 全市
            rec = dict(indicator_name=clean_indicator_name(axis), region="全市", period=period,
                       sheet_name=sheet, confidence="high", generator_type="RULE", source_type="XLSX")
            if vcol is not None and vals.get(vcol) is not None:
                rec["value"], rec["unit"] = vals[vcol], col_unit.get(vcol)
            if gcol is not None and vals.get(gcol) is not None:
                rec["growth_rate"] = vals[gcol]
            if "value" in rec or "growth_rate" in rec:
                out.append(rec)
    return out


def parse_xlsx_workbook(data, doc_title, doc_date=None):
    """解析整个月报卡文件 -> 指标行列表。doc_title 用于推断年份。"""
    if openpyxl is None:
        return []
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    year = infer_year(doc_title, doc_date)
    out = []
    for name in wb.sheetnames:
        sheet = name.strip()
        if sheet.startswith(_SHEET_SKIP):
            continue
        try:
            rows = parse_xlsx_sheet(wb[name], sheet, year)
        except Exception as exc:
            print("  [sheet 跳过] %s -> %s" % (sheet, exc), file=sys.stderr)
            rows = []
        out.extend(rows)
    wb.close()
    return out


def insert_indicator_rows(cur, stat_doc_id, rows):
    sql = ("INSERT IGNORE INTO stat_indicator "
           "(stat_doc_id, period, region, indicator_code, indicator_name, value, unit, "
           "growth_rate, sheet_name, source_type, confidence, generator_type) "
           "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)")
    args = []
    for r in rows:
        args.append((stat_doc_id, r["period"], r["region"],
                     r.get("indicator_code") or re.sub(r"\s+", "", r["indicator_name"]),
                     r["indicator_name"], r.get("value"), r.get("unit"), r.get("growth_rate"),
                     r.get("sheet_name", ""), r.get("source_type"), r.get("confidence", "medium"),
                     r.get("generator_type", "RULE")))
    cur.executemany(sql, args)


def stage2(conn, cur, session, limit, force, dry_run, verbose):
    extra = "" if force else "AND parse_status IN ('DONE')"
    cur.execute("SELECT id, gov_record_id, title, source_url, attachment_url, attachment_name "
                "FROM stat_doc WHERE category='统计月报' %s ORDER BY id" % extra)
    docs = cur.fetchall()
    if limit:
        docs = docs[:limit]
    ok = fail = no_attach = total_ins = 0
    for did, rid, title, surl, att_url, att_name in docs:
        if not att_url:
            no_attach += 1
            if not dry_run:
                cur.execute("UPDATE stat_doc SET parse_status='TEXT_DONE', fail_reason='无附件链接' WHERE id=%s", (did,))
                conn.commit()
            continue
        try:
            resp = session.get(att_url, timeout=TIMEOUT)
            resp.raise_for_status()
            rows = parse_xlsx_workbook(resp.content, title)
        except Exception as exc:
            fail += 1
            print("[xlsx失败] %s -> %s" % (att_url, exc), file=sys.stderr)
            if not dry_run:
                cur.execute("UPDATE stat_doc SET parse_status='XLSX_FAIL', fail_reason=%s WHERE id=%s",
                            (str(exc)[:400], did))
                conn.commit()
            time.sleep(INTERVAL)
            continue
        if dry_run:
            print("[dry] doc=%s (%s) 指标 %d 条 样例: %s"
                  % (did, title[:22], len(rows), json.dumps(rows[:2], ensure_ascii=False)))
        else:
            if rows:
                insert_indicator_rows(cur, did, rows)
                total_ins += len(rows)
            cur.execute("UPDATE stat_doc SET parse_status='XLSX_DONE', fail_reason=NULL WHERE id=%s", (did,))
            conn.commit()
            ok += 1
            if verbose:
                print("[xlsx] %s -> %d 条指标" % (title[:26], len(rows)))
        time.sleep(INTERVAL)
    print("stage2 完成：共 %d 份月报卡，成功 %d / 失败 %d / 无附件 %d，新增指标 %d 条"
          % (len(docs), ok, fail, no_attach, total_ins))


# ============================================================
# 阶段 3：正文指标抽取（规则优先 + LLM 兜底）
# ============================================================
def _find_indicator_value(window):
    """在「增长X%」前 80 字窗口内找 (规范指标名, 数值, 单位)；找不到已知指标时走通用模式。"""
    name, num, unit = "", None, ""
    for alias, mapped in _KNOWN_RULES:
        m = re.search(alias, window)
        if m:
            name = mapped
    um = re.findall(r"%s\s*%s" % (_NUM_RE, _UNIT_RE), window)
    if um:
        num, unit = um[-1]
        num = _to_number(num)
    if name and num is not None:
        return name, num, unit
    # 通用兜底：最近一个「名称+数值+单位」片段
    m = re.search(r"([\u4e00-\u9fa5（）()]{2,10}?)(?:完成|实现|达|为)?\s*%s\s*%s" % (_NUM_RE, _UNIT_RE), window)
    if not m:
        return "", None, ""
    nm = re.sub(r"[，。、,：:]+$", "", m.group(1))
    if len(nm) < 2 or re.search(r"(其中|比|较|年|月|增长|下降|占比|第一|第二|第三|累计)", nm):
        return "", None, ""
    return nm, _to_number(m.group(2)), m.group(3)


def rule_extract_indicators(text, source_type, default_period):
    """规则正则抽取：匹配「指标X数值单位、增长/下降Y%」模式，产出指标行。"""
    out, seen = [], set()
    for line in normalize_text(text or "").split("\n"):
        for m in re.finditer(r"增长(?:了)?\s*(-?\d+(?:\.\d+)?)\s*%", line):
            growth = float(m.group(1))
            name, value, unit = _find_indicator_value(line[max(0, m.start() - 80):m.start()])
            if not name:
                continue
            code = re.sub(r"\s+", "", name)
            key = (code, default_period, "全市")
            if key in seen:
                continue
            seen.add(key)
            out.append(dict(indicator_name=name, indicator_code=code, period=default_period,
                            region="全市", value=value, unit=unit, growth_rate=growth,
                            sheet_name="", source_type=source_type, confidence="medium",
                            generator_type="RULE"))
        for m in re.finditer(r"下降(?:了)?\s*(-?\d+(?:\.\d+)?)\s*%", line):
            growth = -float(m.group(1))
            name, value, unit = _find_indicator_value(line[max(0, m.start() - 80):m.start()])
            if not name:
                continue
            code = re.sub(r"\s+", "", name)
            key = (code, default_period, "全市")
            if key in seen:
                continue
            seen.add(key)
            out.append(dict(indicator_name=name, indicator_code=code, period=default_period,
                            region="全市", value=value, unit=unit, growth_rate=growth,
                            sheet_name="", source_type=source_type, confidence="medium",
                            generator_type="RULE"))
    return out


def llm_extract_indicators(text, source_type, default_period):
    """LLM 兜底抽取：AI_API_KEY 存在时调 DeepSeek 抽 JSON 指标数组；无 key 返回空。"""
    key = (os.environ.get("AI_API_KEY") or "").strip()
    if not key or key in ("your-api-key", "sk-xxx"):
        return []
    model = os.environ.get("AI_MODEL") or "deepseek-chat"
    endpoint = os.environ.get("AI_ENDPOINT") or "https://api.deepseek.com/v1"
    prompt = (
        "你是政府统计数据抽取助手。从下面的统计公报/分析正文中抽取经济指标，输出 JSON 数组，"
        "每项格式：{\"indicator\":\"指标名称\",\"value\":数值,\"unit\":\"单位\","
        "\"growth_rate\":同比增速数值或null,\"period\":\"期间\"}。"
        "只输出 JSON，不要解释。\n\n正文（节选）:\n" + (text or "")[:6000]
    )
    try:
        resp = requests_post(endpoint.rstrip("/") + "/chat/completions",
                             headers={"Authorization": "Bearer " + key,
                                      "Content-Type": "application/json"},
                             json={"model": model, "temperature": 0.2,
                                   "messages": [{"role": "user", "content": prompt}]},
                             timeout=120)
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        items = json.loads(content.strip().lstrip("```json").rstrip("```").strip())
    except Exception as exc:
        print("[LLM 抽取失败] %s" % exc, file=sys.stderr)
        return []
    out = []
    for it in items if isinstance(items, list) else []:
        name = normalize_text(it.get("indicator") or "")
        value = _to_number(it.get("value"))
        growth = _to_number(it.get("growth_rate"))
        if not name or (value is None and growth is None):
            continue
        out.append(dict(indicator_name=name, indicator_code=re.sub(r"\s+", "", name),
                        period=normalize_period(it.get("period"), "") or default_period,
                        region="全市", value=value, unit=it.get("unit") or "",
                        growth_rate=growth, sheet_name="", source_type=source_type,
                        confidence="medium", generator_type="LLM"))
    return out


def stage3(conn, cur, session, categories, limit, force, dry_run, verbose):
    cats = [c for c in categories if c in ("统计公报", "统计分析")]
    if not cats:
        print("stage3 仅处理 统计公报/统计分析")
        return
    marks = ",".join(["%s"] * len(cats))
    extra = "" if force else "AND parse_status IN ('DONE')"
    cur.execute("SELECT id, title, source_url, content, category FROM stat_doc "
                "WHERE category IN (%s) %s ORDER BY id" % (marks, extra), cats)
    docs = cur.fetchall()
    if limit:
        docs = docs[:limit]
    total_ins = done = failed = 0
    for did, title, url, content, cat in docs:
        period = period_from_title(title)
        st = "BULLETIN" if cat == "统计公报" else "ANALYSIS"
        rows = rule_extract_indicators(content or "", st, period)
        gen = "RULE"
        if not rows:
            rows = llm_extract_indicators((content or "")[:6000], st, period)
            gen = "LLM" if rows else "RULE"
        if dry_run:
            print("[dry] doc=%s (%s) %s -> %d 条 样例:%s"
                  % (did, title[:22], gen, len(rows), json.dumps(rows[:2], ensure_ascii=False)))
            continue
        if rows:
            insert_indicator_rows(cur, did, rows)
            total_ins += len(rows)
        cur.execute("UPDATE stat_doc SET parse_status='TEXT_DONE', fail_reason=NULL WHERE id=%s", (did,))
        conn.commit()
        done += 1
        if verbose:
            print("[%s] %s -> %s %d 条" % (cat, title[:26], gen, len(rows)))
    print("stage3 完成：共 %d 篇，处理 %d，新增指标 %d 条" % (len(docs), done, total_ins))


def requests_post(url, **kwargs):
    import requests
    return requests.post(url, **kwargs)


# ============================================================
# CLI
# ============================================================
def parse_args():
    p = argparse.ArgumentParser(description="统计类数据结构化采集管道")
    p.add_argument("--stage", choices=["1", "2", "3", "all"], default="all", help="阶段")
    p.add_argument("--category", default="all",
                   help="tjyb/stjgb/stjfx 逗号分隔，默认全部")
    p.add_argument("--limit", type=int, default=0, help="只处理前 N 条（调试）")
    p.add_argument("--force", action="store_true", help="重跑已处理文档")
    p.add_argument("--dry-run", action="store_true", help="只打印不落库")
    p.add_argument("--verbose", action="store_true", help="逐条打印")
    return p.parse_args()


def main():
    args = parse_args()
    cats = [c.strip() for c in args.category.split(",") if c.strip() in CATEGORY_MAP]
    if args.category != "all" and not cats:
        print("category 无效，可选: tjyb/stjgb/stjfx", file=sys.stderr)
        sys.exit(2)
    if not cats:
        cats = list(CATEGORY_MAP)
    cats = [CATEGORY_MAP[c] for c in cats]  # CLI 键 -> 库内中文类目名
    conn = pymysql.connect(**DB)
    cur = conn.cursor()
    session = make_session()
    stages = ["1", "2", "3"] if args.stage == "all" else [args.stage]
    for st in stages:
        if st == "1":
            stage1(conn, cur, session, cats, args.limit, args.force, args.dry_run, args.verbose)
        elif st == "2":
            stage2(conn, cur, session, args.limit, args.force, args.dry_run, args.verbose)
        else:
            stage3(conn, cur, session, cats, args.limit, args.force, args.dry_run, args.verbose)
    cur.close()
    conn.close()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n已手动中断。", file=sys.stderr)
        sys.exit(130)